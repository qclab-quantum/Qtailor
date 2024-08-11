
import pandas as pd
import  re
import openpyxl
class ExcelUtil:
    @staticmethod
    def read_by_sheet(excel_file):
        # Load the Excel file
        xls = pd.ExcelFile(excel_file)

        # Get a list of all sheet names
        sheet_names = xls.sheet_names

        # Print the list of sheet names
        #print(sheet_names)

        data = {}
        for sheet_name in sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            data.update({sheet_name:df})
        #print(data['qnn'])
        return  sheet_names,data

    @staticmethod
    #2d array to a sheet
    def array2sheet(file_name, sheet_name, data):
        try:
            # Try to load the existing workbook
            workbook = openpyxl.load_workbook(file_name)
        except FileNotFoundError:
            # If the file does not exist, create a new workbook
            workbook = openpyxl.Workbook()

        # If the sheet does not exist, create it
        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(sheet_name)
        else:
            sheet = workbook[sheet_name]

        # Write the 2D array to the sheet
        for row_index, row in enumerate(data):
            for col_index, value in enumerate(row):
                sheet.cell(row=row_index + 1, column=col_index + 1, value=value)

        # Save the workbook
        workbook.save(file_name)

    # Example usage
    data = [
        [1, 2, 3],
        [4, 5, 6],
        [7, 8, 9]
    ]



# def demo():
#     # get data
#     sheets,dfs = ExcelUtil.read_by_sheet('d:/temp.xlsx')
#     # pharse data
#     for sheet in sheets:
#         df = dfs[sheet]
#         circuits = df['circuit']
#         #从字符串中提取出该线路的比特数量 qnn/qnn_indep_qiskit_5.qasm-> 5
#         labels = list(map(lambda x: ''.join(re.findall(r'\d', x)), circuits))
#         print(labels)
#
#         rl = df['rl']
#         qiskit = df['qiskit']
#         mix= df['mix']

# def row2array(sheet):
#     data = []
#     df = dfs[sheet]
#     for index, row in df.iterrows():
#         # print(f"Index: {index}")
#         # print(f"Row:\n{row}\n")
#         temp = []
#         for header in ['ae_30',	'qnn_20',	'su2',	'portfvqe_15',	'pricingcall_21','qtf_20',	'pricingput_17',	]:
#             temp.append(row[header])
#         data.append(temp)
#     return data


if __name__ == '__main__':
    #sheets, dfs = ExcelUtil.read_by_sheet('d:/parallel.xlsx')
    # qtailor = sheets[0]
    # qiskit = sheets[1]
    # q1=row2array(qtailor)
    # q2=row2array(qiskit)

    # Example usage
    data = [
        [1, 2, 3],
        [4, 5, 6],
        [7, 8, 9]
    ]

    ExcelUtil.array2sheet('d:/example.xlsx', 's.a', data)